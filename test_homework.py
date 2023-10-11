import os
from zipfile import ZipFile
import openpyxl
from PyPDF2 import PdfReader
import pandas as pd


def test_txt_file():
    with ZipFile('tmp/testzip.zip') as zf:
        assert zf.read('txt_file.txt').decode('utf-8') == 'txt_file'
        assert zf.getinfo('txt_file.txt').file_size == os.path.getsize('resources/txt_file.txt')


def test_pdf_file():
    with ZipFile('tmp/testzip.zip') as zf:
        archived_pdf = PdfReader(zf.open('pdf_file.pdf'))
        pdf = PdfReader('resources/pdf_file.pdf')
        assert len(archived_pdf.pages) == len(pdf.pages)
        assert 'Пример PDF файла' in archived_pdf.pages[0].extract_text()
        assert zf.getinfo('pdf_file.pdf').file_size == os.path.getsize('resources/pdf_file.pdf')


def test_xls_file():
    with ZipFile('tmp/testzip.zip') as zf:
        archived_book_read = pd.read_excel(zf.open('xls_file.xls'))
        archived_book_excel = pd.ExcelFile(zf.open('xls_file.xls'))
        non_archived_book_read = pd.read_excel('resources/xls_file.xls')
        non_archived_book_excel = pd.ExcelFile('resources/xls_file.xls')
        assert archived_book_read.columns.any() == non_archived_book_read.columns.any() and archived_book_read.columns.all() == non_archived_book_read.columns.all()
        assert archived_book_excel.sheet_names == non_archived_book_excel.sheet_names
        assert archived_book_read.shape[0] == non_archived_book_read.shape[0] and archived_book_read.shape[1] == non_archived_book_read.shape[1]
        assert zf.getinfo('xls_file.xls').file_size == os.path.getsize('resources/xls_file.xls')


def test_xlsx_file():
    with ZipFile('tmp/testzip.zip') as zf:
        archived_book = openpyxl.load_workbook(zf.open('xlsx_file.xlsx')).active
        book = openpyxl.load_workbook('resources/xlsx_file.xlsx').active
        assert archived_book.cell(row=1, column=1).value == book.cell(row=1, column=1).value
        assert archived_book.max_column == book.max_column
        assert archived_book.max_row == book.max_row
        assert zf.getinfo('xlsx_file.xlsx').file_size == os.path.getsize('resources/xlsx_file.xlsx')
